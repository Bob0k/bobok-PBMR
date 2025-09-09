from math import isclose

from openpyxl import Workbook

from constants import *
from parameters import *
from headings import table_head
from cell import Cell

# Initial conditions
temperature = 273.15 + ambienttemperature
incell = Cell(name = "In Cell",
        ch4in = ch4inletpressure * V1 / RT(temperature),
        h2oin = h2oinletpressure * V1 / RT(temperature),
        coin  = 0,
        co2in = 0,
        h2in  = 0,
        volume = V1,
        k1 = 0,
        k2 = 0,
        k3 = 0,
    )
cells = [
    Cell(name = "Main Reactor",
        ch4in = initialpressure * V1 / RT(temperature),
        h2oin = initialpressure * V1 / RT(temperature),
        coin  = initialpressure * V1 / RT(temperature),
        co2in = initialpressure * V1 / RT(temperature),
        h2in  = initialpressure * V1 / RT(temperature),
        volume = V1,
        k1 = f1rate,
        k2 = f2rate,
        k3 = f3rate,
    ),
    Cell(name = "Shift Reactor",
        ch4in = initialpressure * V2 / RT(temperature),
        h2oin = initialpressure * V2 / RT(temperature),
        coin  = initialpressure * V2 / RT(temperature),
        co2in = initialpressure * V2 / RT(temperature),
        h2in  = initialpressure * V2 / RT(temperature),
        volume = V2,
        k1 = s1rate,
        k2 = s2rate,
        k3 = s3rate,
    ),
    ]
burncell = Cell(name = "Burning",
        ch4in = 0,
        h2oin = 0,
        coin  = 0,
        co2in = 0,
        h2in  = 0,
        volume = burncellvolume,
        k1 = 0,
        k2 = 0,
        k3 = 0,
    )
outcells = [
    Cell(name = "Out Cell 1",
        ch4in = 0,
        h2oin = 0,
        coin  = 0,
        co2in = 0,
        h2in  = 0,
        volume = V1h,
        k1 = 0,
        k2 = 0,
        k3 = 0,
         ),
    Cell(name = "Out Cell 2",
        ch4in = 0,
        h2oin = 0,
        coin  = 0,
        co2in = 0,
        h2in  = 0,
        volume = V2h,
        k1 = 0,
        k2 = 0,
        k3 = 0,
         )
           ]
bigoutcell = Cell(name = "Big Out Cell",
        ch4in = 0,
        h2oin = 0,
        coin  = 0,
        co2in = 0,
        h2in  = 0,
        volume = 100000000,
        k1 = 0,
        k2 = 0,
        k3 = 0
                  )
incell.add_neighbour(cells[0], conductivity = incellconductivity)
cells[0].add_neighbour(outcells[0])
cells[1].add_neighbour(outcells[1])
cells[0].add_neighbour(cells[1], conductivity = cell_zero_to_cell_one_conductivity)
cells[1].add_neighbour(burncell, conductivity = cell_one_to_burncell_conductivity)
outcells[0].add_neighbour(bigoutcell, conductivity = outcells_to_outlet_conductivity)
outcells[1].add_neighbour(bigoutcell, conductivity = outcells_to_outlet_conductivity)

lastH2produced = 0

# Excel stuff
wb = Workbook()
data_sheet = wb.active
data_sheet.title = "Data"
# Title writing
row = 1
def excel_write(*args):
    global row
    for i, arg in enumerate(args):
        col = chr(i+97)
        if i > 25:
            col = 'A' + chr(i-26+97)
        data_sheet[col+str(row)] = arg
    row += 1

for head in table_head:
    excel_write(*head)
def write_cell_data(step):
    global lastH2produced
    r1 = []
    r2 = []
    r3 = []
    for cell in cells:
        r1_one, r2_one, r3_one = cell.react_rates(temperature)
        r1.append(r1_one)
        r2.append(r2_one)
        r3.append(r3_one)
    args = [time,
            temperature]
    for i in range(len(cells)):          
        args.append(cells[i].p("ch4", temperature))
        args.append(cells[i].p("h2o", temperature))
        args.append(cells[i].p("co", temperature))
        args.append(cells[i].p("co2", temperature))
        args.append(cells[i].p("h2", temperature))
        args.append(r1[i])
        args.append(r2[i])
        args.append(r3[i])
    args.append(burncell.nu["ch4"])
    args.append(burncell.nu["h2o"])
    args.append(burncell.nu["co"])
    args.append(burncell.nu["co2"])
    args.append(burncell.nu["h2"])
    args.append(outcells[0].p("h2", temperature))
    args.append(outcells[1].p("h2", temperature))
    args.append(outcells[0].nu["h2"])
    args.append(outcells[1].nu["h2"])
    args.append(bigoutcell.nu["h2"])
    args.append((bigoutcell.nu["h2"] - lastH2produced) / step)
    lastH2produced = bigoutcell.nu['h2']
    excel_write(*args)
adaptations = 0
write_cell_data(1)

lasttemperature = temperature
lastcells0ch4 = cells[0].nu['ch4']
try:
    while time < time_max:
        step = initial_step
        adapt_flag = False
        if to_adapt: # Change to 0 to not adapt
            for key in ["ch4","h2o","co","co2","h2"]:
                for i in range(len(cells)):
                    adapt_flag = adapt_flag or (
                        step * abs(cells[i].react_nu_change(temperature)[key])
                        > abs(cells[i].nu[key])
                        )
                    if (
                        step * abs(cells[i].react_nu_change(temperature)[key])
                        > abs(cells[i].nu[key])
                        ):
                        adapt_key = key
            if adapt_flag:
                adaptations += 1
                step = min(
                    [cells[j].get_step(temperature,
                        key = adapt_key,
                        divisor = step_divisor
                        ) for j in range(len(cells))]
                    )
        time += step

        # Logic
        for cell in cells:
            r1r, r2r, r3r = cell.react(step, temperature)
            temperature -= (
                dH('1', temperature) * r1r
                + dH('2', temperature) * r2r
                + dH('3', temperature) * r3r
                ) / totalheatcapacity / overcoolingfactor
            if temperature < 0:
                for i in ["temperature",
                      "dH('1', temperature)",
                      "r1r",
                      "dH('2', temperature)",
                      "r2r",
                      "dH('3', temperature)",
                      "r3r",
                      "totalheatcapacity"]:
                    print(i, eval(i))
                exit()
            
            
        for cell in cells + outcells:
            cell.give(step, temperature)
        incell.give(step, temperature)
        incell.nu['ch4'] = ch4inletpressure * V1 / RT(temperature)
        incell.nu['h2o'] = h2oinletpressure * V1 / RT(temperature)

        temperature += burncell.burn(percent = burnpercent)
        
        # Writing
        #write_cell_data()
        write_counter += 1
        if write_counter == write_trigger:
            write_cell_data(step)
            write_counter = 0
        burncell.nu['h2o'] = 0
        burncell.nu['co2'] = 0

        if time / time_max > print_threshold:
            print(f"{int(print_threshold*100)}% done")
            print_threshold += 0.1
        break_flag = True
        for cell in cells:
            break_flag = (
                break_flag
                and not adapt_flag
                and isclose(lasttemperature, temperature)
                and isclose(lastcells0ch4, cells[0].nu['ch4'])
                ) or temperature > 1000
        if break_flag:
            print("Broken")
            break
        else:
            lasttemperature = temperature
            lastcells0ch4 = cells[0].nu['ch4']
            
except Exception as e:
    print(e)
finally:
    #write_cell_data(step)
    print(f"100% done")
    print("Saving in progress")
    parameters_sheet = wb.create_sheet("Parameters")
    row = 1
    col = 1
    for d in give_parameters_dicts():
        for key in d:
            parameters_sheet[chr(col+96) + str(row)] = key
            parameters_sheet[chr(col+97) + str(row)] = d[key]
            row += 1
        row = 1
        col += 2
##    from openpyxl.chart import LineChart, Reference
##    from openpyxl.chart.shapes import GraphicalProperties
##    charts_sheet = wb.create_sheet("Charts")
##    charts_count = 0
##    for i in range(len(table_head[0])-1):
##        chart = LineChart()
##        data = Reference(data_sheet, min_col = 1, min_row = 1)
##        chart.add_data(data, titles_from_data = True)
##        
##        chart.graphical_properties = GraphicalProperties()
##        chart.graphical_properties.line.noFill = True
##        chart.graphical_properties.line.prstDash = None
##        
##        charts_sheet.add_chart(chart, "A" + 14 * str(charts_count))
##        charts_count += 1
##        print(charts_count)
    from datetime import datetime
    filename = (
        "outputs/"
        + control_sum()
        + str(datetime.now()).replace(":"," ")
        )
    wb.save(filename + ".xlsx")
    print(f"Saved as {filename}")
    print(f"{adaptations} adaptations performed")
    input("Press Enter to exit")
